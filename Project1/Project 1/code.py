"""
P&ID Legend Analyzer  v7  ─  All Fixes Applied
================================================

CHANGES FROM v6
───────────────
  Fix #1  : YOLO-World with P&ID class prompts (replaces generic COCO model)
  Fix #2  : Separate CLIP image/text scoring (no mixed-vector queries)
  Fix #3  : Word-boundary alias matching, longest-first, no greedy break
  Fix #4  : Aspect-ratio VLM crop with minimum-size padding
  Fix #5  : Lateral/omnidirectional section assignment
  Fix #6  : Improved inpainting with symbol-edge protection
  Fix #7  : Alias-aware unmatched label comparison
  Fix #8  : Max-width token grouping to prevent cross-column merges
  Fix #9  : Raised YOLO confidence threshold (0.15 → 0.25)
  Fix #10 : Set-based deduplication in KB retrieval
  Fix #11 : Fixed save_excel parameter passing
  Fix #12 : Grid-aware symbol-label pairing within sections

PIPELINE (per legend page)
───────────────────────────
  1.  OCR full page → extract every text token + confidence + bbox
  2.  Export raw OCR to Excel (Sheet: "OCR Raw Text")
  3.  Group tokens → label groups (union-find, same row + h-gap, max-width cap)
  4.  Classify groups → header / label / noise (semantic rules)
  5.  Erase ALL text from image (inpaint with symbol protection) → clean image
  6.  YOLO-World + manual multi-scale slicing on clean image
  7.  Contour fallback (fine + medium) on clean image
  8.  NMS across all sources → final detection list
  9.  Section assignment (omnidirectional header matching)
  10. Grid-aware symbol↔label pairing within sections
  11. Per-symbol: crop from ORIGINAL image, run CLIP embed
  12. Retrieve top-k from KB (separate image & text cosine scores)
  13. VLM confirmation (GPT-4V or local LLaVA) with properly sized crop
  14. Write Excel: 4 sheets (all symbols, section summary, OCR text, unmatched)
"""

import io, os, re, math, base64, json, time, argparse, sys
import cv2, numpy as np
from pathlib import Path
from PIL import Image as PILImage
from collections import Counter, defaultdict

PILImage.MAX_IMAGE_PIXELS = None

# PyMuPDF replaces pdf2image / poppler
import fitz  # pip install pymupdf

# EasyOCR is optional — we prefer PDF text layer extraction
try:
    import easyocr
    EASYOCR_AVAILABLE = True
except ImportError:
    EASYOCR_AVAILABLE = False

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import (Font, PatternFill, Alignment,
                               Border, Side)
from openpyxl.chart import BarChart, Reference

# ═══════════════════════════════════════════════════════════════════════
# CONFIGURATION  — defaults (overridden at runtime by CLI args + JSON DB)
# ═══════════════════════════════════════════════════════════════════════
PDF_DPI           = 150
YOLO_MODEL_PATH   = "yolov8s-world.pt"
CONFIDENCE_THRES  = 0.25

USE_VLM           = False
OPENAI_API_KEY    = os.getenv("OPENAI_API_KEY", "")
VLM_MODEL         = "gpt-4o"
VLM_TOP_K         = 3
VLM_RETRY         = 2
VLM_MIN_CROP_PX   = 256

USE_CLIP          = False
CLIP_MODEL_NAME   = "openai/clip-vit-base-patch32"
CLIP_IMG_SCORE_W  = 0.60
CLIP_TXT_SCORE_W  = 0.40

SLICE_SIZES       = [320, 512, 768]
OVERLAP_RATIO     = 0.30
NMS_IOU_THRESH    = 0.35
WBF_IOU_THRESH    = 0.50

SYM_MIN_AREA      = 120
SYM_MAX_AREA      = 320_000
SYM_MIN_DIM       = 8

OCR_MIN_CONF      = 0.28
OCR_MIN_LEN       = 2

H_GAP_PX          = 40
V_GAP_PX          = 10
ROW_TOL_PX        = 18
MAX_GROUP_WIDTH_PX = 380

HEADER_MIN_HEIGHT  = 12
HEADER_UC_FRAC     = 0.78

ASSOC_RIGHT_MAX   = 500
ASSOC_ANY_MAX     = 260

INPAINT_PAD       = 6
INPAINT_RADIUS    = 5

THUMB_W, THUMB_H  = 88, 66

# Runtime-populated from JSON database (see _load_config)
PID_KNOWLEDGE_BASE = []
PID_YOLO_CLASSES   = []

# Alias lookup — populated at runtime by _load_config() from the --database JSON
# See symbols_db.json for the full knowledge base.
_ALIAS_ENTRIES = []   # list of (alias_lower, canonical_name, alias_len)
_ALIAS_LOOKUP  = {}   # alias_lower → canonical_name

# ═══════════════════════════════════════════════════════════════════════
# CLIP MODEL (lazy-loaded)
# ═══════════════════════════════════════════════════════════════════════
_clip_model = None
_clip_processor = None
_kb_vectors = None   # shape (N, D)  — text embeddings of KB descriptions
_kb_names   = None   # list of N canonical names

def _load_clip():
    global _clip_model, _clip_processor
    if _clip_model is not None: return
    try:
        from transformers import CLIPModel, CLIPProcessor
        print("  [CLIP] Loading model …")
        _clip_model     = CLIPModel.from_pretrained(CLIP_MODEL_NAME)
        _clip_processor = CLIPProcessor.from_pretrained(CLIP_MODEL_NAME)
        _clip_model.eval()
        print("  [CLIP] Ready.")
    except ImportError:
        print("  [CLIP] transformers not installed → falling back to OCR-only matching.")
        global USE_CLIP
        USE_CLIP = False

def _clip_embed_image(bgr: np.ndarray) -> np.ndarray:
    """Return normalised CLIP image embedding (numpy 1D array)."""
    import torch
    rgb = cv2.cvtColor(bgr, cv2.COLOR_BGR2RGB) if bgr.ndim == 3 else cv2.cvtColor(bgr, cv2.COLOR_GRAY2RGB)
    pil = PILImage.fromarray(rgb)
    inputs = _clip_processor(images=pil, return_tensors="pt")
    with torch.no_grad():
        feat = _clip_model.get_image_features(**inputs)
    v = feat.numpy()[0]
    return v / (np.linalg.norm(v) + 1e-8)

def _clip_embed_text(text: str) -> np.ndarray:
    """Return normalised CLIP text embedding."""
    import torch
    inputs = _clip_processor(text=[text], return_tensors="pt", padding=True, truncation=True)
    with torch.no_grad():
        feat = _clip_model.get_text_features(**inputs)
    v = feat.numpy()[0]
    return v / (np.linalg.norm(v) + 1e-8)

def build_kb_vectors():
    """Pre-compute text embeddings for every KB entry (once at startup)."""
    global _kb_vectors, _kb_names
    if not USE_CLIP: return
    _load_clip()
    if not USE_CLIP: return
    print("  [KB] Building knowledge-base vectors …")
    vecs, names = [], []
    for entry in PID_KNOWLEDGE_BASE:
        # Use CLIP-friendly prompt format for text embeddings
        query = f"A P&ID engineering drawing symbol of a {entry['name']}. " \
                f"{entry['shape_desc']}. Section: {entry['section']}."
        vecs.append(_clip_embed_text(query))
        names.append(entry["name"])
    _kb_vectors = np.stack(vecs)
    _kb_names   = names
    print(f"  [KB] {len(_kb_names)} entries embedded.")


# ═══════════════════════════════════════════════════════════════════════
# FIX #2 + #3 + #10: Improved KB Retrieval
#   - Separate image & text cosine scores (don't mix vectors)
#   - Word-boundary alias matching with longest-first priority
#   - Set-based dedup
# ═══════════════════════════════════════════════════════════════════════
def _word_boundary_match(alias: str, text: str) -> bool:
    """
    FIX #3: Check if alias appears in text at word boundaries.
    Prevents 'gate' matching inside 'investigate' or 'pvc' inside 'cpvc'.
    """
    pattern = r'(?<![a-zA-Z])' + re.escape(alias) + r'(?![a-zA-Z])'
    return bool(re.search(pattern, text, re.IGNORECASE))


def kb_retrieve(crop_bgr: np.ndarray, ocr_nearby: str, top_k=5) -> list:
    """
    Retrieve top-k KB entries for a symbol crop.
    Returns list of (name, score) tuples, best first.

    FIX #2: Compute image→KB and text→KB similarity SEPARATELY,
            then combine the scores (not the vectors).
    FIX #3: Word-boundary alias matching, longest-first, collect ALL matches.
    FIX #10: Use set for dedup.
    """
    candidates = []
    seen_names = set()   # FIX #10: set-based dedup
    ocr_low = ocr_nearby.lower().strip()

    # ── FIX #3: Alias matching — longest-first, word-boundary, collect all ──
    alias_matches = []
    for alias, canonical, alen in _ALIAS_ENTRIES:
        if alen < 3:
            continue  # skip dangerously short aliases (nc, no, lo, lc, etc.)
        if _word_boundary_match(alias, ocr_low):
            if canonical not in seen_names:
                # Longer alias = higher confidence (0.85–0.95, not a hard 1.0)
                conf = min(0.95, 0.80 + 0.01 * alen)
                alias_matches.append((canonical, conf))
                seen_names.add(canonical)
    # Take the best alias match(es) — they are already longest-first
    candidates.extend(alias_matches[:2])  # keep at most 2 alias hits

    # ── FIX #2: Separate CLIP image/text scoring ─────────────────────────
    if USE_CLIP and _kb_vectors is not None and crop_bgr is not None and crop_bgr.size > 0:
        # Image → KB cosine similarities
        img_vec  = _clip_embed_image(crop_bgr)
        img_sims = (_kb_vectors @ img_vec)  # shape (N,)

        # Text → KB cosine similarities (if OCR text available)
        if ocr_low.strip():
            txt_vec  = _clip_embed_text(ocr_nearby)
            txt_sims = (_kb_vectors @ txt_vec)  # shape (N,)
        else:
            txt_sims = np.zeros_like(img_sims)

        # Combine SCORES (not vectors!) with weights
        combined = CLIP_IMG_SCORE_W * img_sims + CLIP_TXT_SCORE_W * txt_sims
        ranked   = sorted(enumerate(combined.tolist()), key=lambda x: -x[1])

        for idx, score in ranked[:top_k]:
            name = _kb_names[idx]
            if name not in seen_names:
                candidates.append((name, round(score, 3)))
                seen_names.add(name)
    else:
        # CLIP disabled: fill with first few KB entries after aliases
        for entry in PID_KNOWLEDGE_BASE[:top_k]:
            if entry["name"] not in seen_names:
                candidates.append((entry["name"], 0.0))
                seen_names.add(entry["name"])

    return candidates[:top_k]


# ═══════════════════════════════════════════════════════════════════════
# FIX #4: VLM CONFIRMATION with proper crop sizing
# ═══════════════════════════════════════════════════════════════════════
def _prepare_vlm_crop(crop_bgr: np.ndarray) -> np.ndarray:
    """
    FIX #4: Resize crop maintaining aspect ratio, pad to minimum size.
    - Never downscale below VLM_MIN_CROP_PX on any dimension
    - Pad small crops with white border for context
    - Maintain aspect ratio throughout
    """
    if crop_bgr is None or crop_bgr.size == 0:
        return np.ones((VLM_MIN_CROP_PX, VLM_MIN_CROP_PX, 3), dtype=np.uint8) * 220

    h, w = crop_bgr.shape[:2]

    # Upscale small crops to minimum size while keeping aspect ratio
    if h < VLM_MIN_CROP_PX or w < VLM_MIN_CROP_PX:
        scale = max(VLM_MIN_CROP_PX / h, VLM_MIN_CROP_PX / w)
        new_w, new_h = int(w * scale), int(h * scale)
        resized = cv2.resize(crop_bgr, (new_w, new_h), interpolation=cv2.INTER_CUBIC)
    elif h > 512 or w > 512:
        # Downscale very large crops (equipment) to max 512 maintaining ratio
        scale = min(512 / h, 512 / w)
        new_w, new_h = int(w * scale), int(h * scale)
        resized = cv2.resize(crop_bgr, (new_w, new_h), interpolation=cv2.INTER_AREA)
    else:
        resized = crop_bgr.copy()

    # Pad to square with white border for clean VLM input
    rh, rw = resized.shape[:2]
    target = max(rh, rw, VLM_MIN_CROP_PX)
    canvas = np.ones((target, target, 3), dtype=np.uint8) * 255
    y_off = (target - rh) // 2
    x_off = (target - rw) // 2
    canvas[y_off:y_off+rh, x_off:x_off+rw] = resized
    return canvas


def vlm_confirm(crop_bgr: np.ndarray, ocr_nearby: str,
                 candidates: list, section: str) -> dict:
    """
    Send crop + context to GPT-4V/GPT-4o → get structured JSON back.
    Returns dict: {name, standard_ref, confidence, reasoning}
    Falls back gracefully if API fails.
    """
    if not USE_VLM or not OPENAI_API_KEY:
        if candidates:
            return {"name": candidates[0][0], "standard_ref": "",
                    "confidence": float(candidates[0][1]), "reasoning": "KB retrieval only"}
        return {"name": "Unknown Symbol", "standard_ref": "",
                "confidence": 0.0, "reasoning": "No candidates"}

    try:
        import openai
        openai.api_key = OPENAI_API_KEY
    except ImportError:
        print("  [VLM] openai not installed → using KB only.")
        if candidates:
            return {"name": candidates[0][0], "standard_ref": "",
                    "confidence": float(candidates[0][1]), "reasoning": "KB retrieval only"}
        return {"name": "Unknown Symbol", "standard_ref": "", "confidence": 0.0, "reasoning": "No API"}

    # FIX #4: Proper crop preparation
    prepared = _prepare_vlm_crop(crop_bgr)
    _, buf = cv2.imencode(".png", prepared)
    b64    = base64.b64encode(buf.tobytes()).decode()

    cand_str = ", ".join(f'"{n}" (score:{s:.2f})' for n, s in candidates)
    prompt = f"""You are an expert P&ID (Process & Instrumentation Diagram) symbol recognition system using ISA 5.1 and IEC 60617 standards.

I will show you a cropped symbol from a P&ID legend sheet.

Context:
- Section on the legend: "{section}"
- Nearby OCR text extracted from the image: "{ocr_nearby}"
- Top candidates from vector search: {cand_str}

Instructions:
1. Look at the symbol shape carefully.
2. Consider the nearby OCR text — if it clearly matches a candidate name, weight that heavily.
3. Use your P&ID engineering knowledge to confirm the best match.
4. If the OCR text and the shape both agree on a candidate, set confidence > 0.85.
5. If shape matches but text is ambiguous, set confidence 0.5–0.75.
6. If nothing matches well, return "Unknown Symbol" with confidence < 0.3.

Reply ONLY with valid JSON, no markdown, no explanation outside the JSON:
{{"name": "exact canonical symbol name", "standard_ref": "e.g. ISA 5.1 Fig 3-2 or empty string", "confidence": 0.0, "reasoning": "brief 1-sentence reason"}}"""

    for attempt in range(VLM_RETRY + 1):
        try:
            import openai
            response = openai.chat.completions.create(
                model=VLM_MODEL,
                messages=[{
                    "role": "user",
                    "content": [
                        {"type": "image_url",
                         "image_url": {"url": f"data:image/png;base64,{b64}",
                                       "detail": "high"}},
                        {"type": "text", "text": prompt}
                    ]
                }],
                max_tokens=200,
                temperature=0.1,
            )
            raw = response.choices[0].message.content.strip()
            raw = re.sub(r"```json|```", "", raw).strip()
            result = json.loads(raw)
            for k in ["name", "confidence", "reasoning"]:
                if k not in result:
                    result[k] = ("Unknown" if k == "name" else
                                 0.0 if k == "confidence" else "")
            result.setdefault("standard_ref", "")
            return result
        except Exception as e:
            if attempt < VLM_RETRY:
                time.sleep(1.5)
            else:
                print(f"  [VLM] API failed: {e}")
                if candidates:
                    return {"name": candidates[0][0], "standard_ref": "",
                            "confidence": float(candidates[0][1]),
                            "reasoning": f"VLM error: {e}"}
                return {"name": "Unknown Symbol", "standard_ref": "",
                        "confidence": 0.0, "reasoning": str(e)}


# ═══════════════════════════════════════════════════════════════════════
# STANDARD UTILITIES
# ═══════════════════════════════════════════════════════════════════════
def setup_dirs(path):
    p = Path(path); p.mkdir(parents=True, exist_ok=True); return p

def pdf_to_image(pdf_path, page, dpi, out_dir, prefix="legend"):
    """Render a single PDF page to a PNG using PyMuPDF (no poppler needed)."""
    doc = fitz.open(pdf_path)
    if page < 1 or page > doc.page_count:
        raise ValueError(f"Page {page} out of range (doc has {doc.page_count} pages)")
    mat = fitz.Matrix(dpi / 72.0, dpi / 72.0)
    pix = doc[page - 1].get_pixmap(matrix=mat, colorspace=fitz.csRGB)
    doc.close()
    p = str(out_dir / f"{prefix}_p{page}.png")
    pix.save(p)
    print(f"    Rendered page {page} @ {dpi} DPI → {Path(p).name}")
    return p

def preprocess_ocr(gray):
    clahe    = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
    enhanced = clahe.apply(gray)
    denoised = cv2.fastNlMeansDenoising(enhanced, h=8)
    return cv2.adaptiveThreshold(denoised, 255,
        cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 25, 8)

def run_ocr_raw(image_path, reader, pdf_path=None, page_num=None, scale=1.0):
    """
    Extract text tokens. Prefers the PDF embedded text layer (zero OCR error).
    Falls back to EasyOCR when no PDF path is supplied or the layer is empty.
    """
    _bad = {"", "|", "_", "-", ".", "~", ",", "!", "?", "/", "\\"}

    # ── Primary: PDF text layer ──────────────────────────────────────────
    if pdf_path and page_num:
        try:
            doc = fitz.open(pdf_path)
            pg  = doc[page_num - 1]
            words = pg.get_text("words")   # list of (x0,y0,x1,y1,text,block,line,word)
            doc.close()
            if words:
                toks = []
                for w in words:
                    x1, y1, x2, y2, text = w[0], w[1], w[2], w[3], w[4]
                    text = text.strip()
                    if len(text) < OCR_MIN_LEN or text in _bad:
                        continue
                    # Scale from PDF pts (72 dpi base) to rendered image pixels
                    x1s = int(x1 * scale); y1s = int(y1 * scale)
                    x2s = int(x2 * scale); y2s = int(y2 * scale)
                    toks.append({
                        "text": text, "confidence": 1.0,
                        "x1": x1s, "y1": y1s, "x2": x2s, "y2": y2s,
                        "cx": (x1s + x2s) // 2, "cy": (y1s + y2s) // 2,
                        "height": max(1, y2s - y1s),
                    })
                print(f"    PDF text layer: {len(toks)} tokens (scale={scale:.3f})")
                return toks
        except Exception as e:
            print(f"    PDF text extraction failed ({e}), falling back to OCR")

    # ── Fallback: EasyOCR ────────────────────────────────────────────────
    if not EASYOCR_AVAILABLE or reader is None:
        print("    WARNING: no OCR available and no PDF text layer extracted")
        return []
    gray  = cv2.imread(image_path, cv2.IMREAD_GRAYSCALE)
    proc  = preprocess_ocr(gray)
    raw   = reader.readtext(proc, detail=1, paragraph=False,
                             min_size=6, text_threshold=0.45,
                             low_text=0.30, link_threshold=0.30)
    toks  = []
    for (box, text, conf) in raw:
        text = text.strip().strip("|_-~•·")
        if conf < OCR_MIN_CONF or len(text) < OCR_MIN_LEN or text in _bad:
            continue
        xs = [p[0] for p in box]; ys = [p[1] for p in box]
        x1, y1, x2, y2 = int(min(xs)), int(min(ys)), int(max(xs)), int(max(ys))
        toks.append({"text": text, "confidence": round(float(conf), 3),
                     "x1": x1, "y1": y1, "x2": x2, "y2": y2,
                     "cx": (x1+x2)//2, "cy": (y1+y2)//2,
                     "height": y2-y1})
    print(f"    OCR: {len(toks)} raw tokens")
    return toks


# ═══════════════════════════════════════════════════════════════════════
# FIX #8: Token grouping with max-width constraint
# ═══════════════════════════════════════════════════════════════════════
def group_tokens(tokens):
    """
    Group nearby OCR tokens into label groups using union-find.
    FIX #8: Added MAX_GROUP_WIDTH_PX to prevent merging across columns.
            Reduced H_GAP_PX from 55 → 40 px.
    """
    if not tokens: return []
    n = len(tokens)
    parent = list(range(n))

    def find(i):
        while parent[i] != i:
            parent[i] = parent[parent[i]]
            i = parent[i]
        return i

    def union(i, j):
        parent[find(i)] = find(j)

    # Track group bounding boxes to enforce max-width
    group_x1 = {i: tokens[i]["x1"] for i in range(n)}
    group_x2 = {i: tokens[i]["x2"] for i in range(n)}

    idx = sorted(range(n), key=lambda i: (tokens[i]["cy"], tokens[i]["cx"]))
    for pi, i in enumerate(idx):
        for pj in range(pi+1, n):
            j = idx[pj]
            a, b = tokens[i], tokens[j]
            if b["y1"] - a["y2"] > V_GAP_PX * 5:
                break
            if (abs(a["cy"] - b["cy"]) <= ROW_TOL_PX and
                not(a["y2"] + V_GAP_PX < b["y1"] or b["y2"] + V_GAP_PX < a["y1"]) and
                max(a["x1"], b["x1"]) - min(a["x2"], b["x2"]) < H_GAP_PX):
                # FIX #8: Check if merging would exceed max group width
                ri, rj = find(i), find(j)
                merged_x1 = min(group_x1.get(ri, a["x1"]), group_x1.get(rj, b["x1"]))
                merged_x2 = max(group_x2.get(ri, a["x2"]), group_x2.get(rj, b["x2"]))
                if merged_x2 - merged_x1 > MAX_GROUP_WIDTH_PX:
                    continue  # skip merge — would span too wide
                union(i, j)
                new_root = find(i)
                group_x1[new_root] = merged_x1
                group_x2[new_root] = merged_x2

    groups = defaultdict(list)
    for i in range(n):
        groups[find(i)].append(tokens[i])
    result = []
    for members in groups.values():
        members.sort(key=lambda t: t["x1"])
        text = " ".join(m["text"] for m in members)
        conf = round(sum(m["confidence"] for m in members) / len(members), 3)
        x1 = min(m["x1"] for m in members); y1 = min(m["y1"] for m in members)
        x2 = max(m["x2"] for m in members); y2 = max(m["y2"] for m in members)
        result.append({"text": text, "confidence": conf,
                        "x1": x1, "y1": y1, "x2": x2, "y2": y2,
                        "cx": (x1+x2)//2, "cy": (y1+y2)//2,
                        "token_count": len(members),
                        "avg_height": sum(m["height"] for m in members) / len(members)})
    print(f"    Grouped: {len(result)} label groups")
    return result


_SECTION_KEYWORDS = {
    # ── P&ID sections (original) ────────────────────────────────────────
    "LINE SYMBOLS", "CONTROL VALVE SYMBOLS", "INSTRUMENT SYMBOLS",
    "PIPING SYMBOLS", "INSULATION SYMBOLS", "VALVE SYMBOLS",
    "RELAY FUNCTION", "DESIGNATION SIGNAL", "INSTRUMENT DETAILS",
    "P&ID REPRESENTATION", "PUMPS AND COMPRESSOR", "VESSELS",
    "MISCELLANEOUS ITEMS", "COLUMN AND HEAT EXCHANGERS",
    "BLOCK/DESIGNATIONS",
    # ── Electrical legend: full exact section-header strings ────────────
    "ELECTRICAL LEGEND",
    "SWITCHING CONTROLS",
    "DISTRIBUTION AND EQUIPMENT",
    "DISTRIBUTION AND",          # split partial (grouper may split multi-word header)
    "POWER DEVICES",
    "REFERENCE SYMBOLS",
    "FIRE ALARM DEVICES",
    "ADDRESSABLE CONTROL MODULE",
    "SENSORS", "SIGNAL DEVICES",
    "GROUNDING SYSTEM",
    "ELECTRICAL POWER MONITORING SYSTEM",
    "ELECTRICAL POWER",          # split partial
    "MONITORING SYSTEM",         # split partial
    "POWER LOAD BALANCING SYSTEM",
    "DYNAMIC LOAD",              # split partial
    "BALANCING SYSTEM",          # split partial
    "ABBREVIATIONS", "GENERAL NOTES",
    "OUTLET MOUNTING HEIGHTS",
    "OUTLET MOUNTING",           # split partial
    "CABLE BUS SPECS",
    "ELECTRICAL EQUIPMENT NAMING CONVENTION LEGEND",
    "ELECTRICAL EQUIPMENT",      # split partial
    "NAMING CONVENTION",         # split partial
    # ── Title-block / stamp text ─────────────────────────────────────────
    "AHJ STAMP", "CONFIDENTIAL", "PERMIT",
    # "LIGHTING" re-added: safe now that we use exact match (not substring).
    "LIGHTING",
    # "CONTROL" still excluded: it appears as part of "SWITCHING CONTROLS" in the
    # legend — adding it would steal symbols from that section.
}

def classify_text_groups(groups):
    # Scale minimum header height from 12pt reference to current DPI
    # (12pt at 72 DPI = 12px; at 150 DPI = 25px; at 300 DPI = 50px)
    _hdr_min_h = max(HEADER_MIN_HEIGHT, int(HEADER_MIN_HEIGHT * PDF_DPI / 72))

    classified = []
    for grp in groups:
        txt = grp["text"].strip()
        upper_frac = sum(1 for c in txt if c.isupper()) / max(len(txt), 1)
        # Exact-match only: txt must equal one of the known section header strings.
        # Substring matching caused false positives (e.g. "CONTROL" inside
        # "FIRE ALARM CONTROL PANEL AND ASSOCIATED").
        is_known_kw = txt.upper() in _SECTION_KEYWORDS

        word_count = len(txt.split())

        # Known section keyword (exact match) → header
        if is_known_kw:
            is_hdr = True
        else:
            # Non-keyword: only promote very short all-caps text to header
            # (1-2 words). Also require Y ≥ title-band limit so page-title
            # words ("ELECTRICAL", "LEGEND" at Y≈117px) are never promoted to
            # section headers. Real section headers start at Y≈186px.
            # Threshold = 80 pts × scale (167px at 150 DPI).
            _title_band_px = int(80 * PDF_DPI / 72)
            is_hdr = (upper_frac >= HEADER_UC_FRAC
                      and grp["avg_height"] >= _hdr_min_h
                      and 2 <= len(txt) <= 20
                      and word_count <= 2
                      and grp.get("y1", 0) >= _title_band_px)

        # Vertical title-block text has an extremely narrow bounding box (≈17 px wide)
        # because the characters are stacked sideways. Exclude these before header test.
        is_narrow_vertical = (grp["x2"] - grp["x1"]) < 30

        # Page-title band: text at the very top of the page (Y1 < ~167 px at 150 DPI)
        # that is NOT a known section keyword is either the drawing title or notes text.
        # Mark as noise so it is never paired as a symbol label.
        _title_band_px = int(80 * PDF_DPI / 72)
        is_title_band_noise = (grp.get("y1", 0) < _title_band_px and not is_known_kw)

        is_noise = (len(txt) < 2
                    or bool(re.fullmatch(r'[\d\s\.\-\,]+', txt))
                    or len(txt) == 1
                    or is_narrow_vertical
                    or is_title_band_noise)
        grp["kind"] = "noise" if is_noise else ("header" if is_hdr else "label")
        classified.append(grp)
    cnts = Counter(g["kind"] for g in classified)
    print(f"    Classified: {cnts}  (hdr_min_h={_hdr_min_h}px)")
    return classified


# ═══════════════════════════════════════════════════════════════════════
# FIX #6: Improved text erasure with symbol-edge protection
# ═══════════════════════════════════════════════════════════════════════
def erase_text(image_path, tokens, out_dir, symbol_bboxes=None):
    """
    Inpaint text regions to create a clean image for symbol detection.
    FIX #6: Use smaller inpaint radius near detected edges to avoid
            smearing symbol contours that touch text bboxes.
    """
    img = cv2.imread(image_path)
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    mask = np.zeros(img.shape[:2], dtype=np.uint8)
    pad = INPAINT_PAD

    # Build edge map to protect symbol edges during inpainting
    edges = cv2.Canny(gray, 50, 150)
    edge_dilated = cv2.dilate(edges, np.ones((3, 3), dtype=np.uint8), iterations=1)

    for tok in tokens:
        x1 = max(0, tok["x1"] - pad); y1 = max(0, tok["y1"] - pad)
        x2 = min(img.shape[1]-1, tok["x2"] + pad)
        y2 = min(img.shape[0]-1, tok["y2"] + pad)
        mask[y1:y2, x1:x2] = 255

    # FIX #6: Subtract strong edge pixels from the inpaint mask
    # This prevents inpainting from destroying symbol lines that
    # happen to touch text bounding boxes
    mask_protected = cv2.subtract(mask, edge_dilated)
    # Re-add the core text region (without padding) to ensure text IS erased
    for tok in tokens:
        mask_protected[tok["y1"]:tok["y2"], tok["x1"]:tok["x2"]] = 255

    k = cv2.getStructuringElement(cv2.MORPH_RECT, (INPAINT_PAD, INPAINT_PAD))
    mask_final = cv2.dilate(mask_protected, k, iterations=1)
    clean = cv2.inpaint(img, mask_final, INPAINT_RADIUS, cv2.INPAINT_TELEA)

    stem = Path(image_path).stem
    out = str(out_dir / f"{stem}_clean.png")
    cv2.imwrite(out, clean)
    print(f"    Text erased → {Path(out).name}")
    return out


# ═══════════════════════════════════════════════════════════════════════
# DETECTION UTILITIES
# ═══════════════════════════════════════════════════════════════════════
def iou(a, b):
    ax1, ay1, ax2, ay2 = a; bx1, by1, bx2, by2 = b
    ix1, iy1 = max(ax1, bx1), max(ay1, by1)
    ix2, iy2 = min(ax2, bx2), min(ay2, by2)
    inter = max(0, ix2-ix1) * max(0, iy2-iy1)
    union = (ax2-ax1)*(ay2-ay1) + (bx2-bx1)*(by2-by1) - inter
    return inter / union if union > 0 else 0.0

def wbf(boxes_list, scores_list, iou_thr=WBF_IOU_THRESH):
    ab = [b for bl in boxes_list for b in bl]
    as_ = [s for sl in scores_list for s in sl]
    if not ab: return [], []
    order = sorted(range(len(as_)), key=lambda i: -as_[i])
    merged = []; ms = []; used = [False]*len(ab)
    for i in order:
        if used[i]: continue
        cb = [ab[i]]; cs = [as_[i]]; used[i] = True
        for j in order:
            if used[j]: continue
            if iou(ab[i], ab[j]) > iou_thr:
                cb.append(ab[j]); cs.append(as_[j]); used[j] = True
        ws = np.array(cs); ws /= ws.sum()
        bx = np.array(cb, dtype=float)
        merged.append((int((bx[:, 0]*ws).sum()), int((bx[:, 1]*ws).sum()),
                       int((bx[:, 2]*ws).sum()), int((bx[:, 3]*ws).sum())))
        ms.append(float(max(cs)))
    return merged, ms


# ═══════════════════════════════════════════════════════════════════════
# FIX #1: YOLO-World with P&ID class prompts + manual sliced prediction
# Replaces run_yolo_sahi() which used a generic COCO model via SAHI
# ═══════════════════════════════════════════════════════════════════════
def _load_yolo_world(model_path):
    """
    Load YOLO-World model and set P&ID-specific class prompts.
    Falls back to standard YOLO + SAHI if YOLO-World fails.
    """
    try:
        from ultralytics import YOLO
        model = YOLO(model_path)
        # YOLO-World specific: set open-vocabulary classes
        if hasattr(model, 'set_classes'):
            model.set_classes(PID_YOLO_CLASSES)
            print(f"  [YOLO-World] Loaded with {len(PID_YOLO_CLASSES)} P&ID classes")
        else:
            print(f"  [YOLO] Loaded standard model (no set_classes support)")
        return model, "yolo_world"
    except Exception as e:
        print(f"  [YOLO] ultralytics load failed: {e}")
        print(f"  [YOLO] Falling back to SAHI + standard YOLO")
        try:
            from sahi import AutoDetectionModel
            model = AutoDetectionModel.from_pretrained(
                model_type="yolov8", model_path=model_path,
                confidence_threshold=CONFIDENCE_THRES, device="cpu")
            return model, "sahi"
        except Exception as e2:
            print(f"  [YOLO] SAHI fallback also failed: {e2}")
            return None, "none"


def run_yolo_world_sliced(image_path, yolo_model):
    """
    FIX #1: Run YOLO-World with manual multi-scale sliced prediction.
    Uses the P&ID class prompts set during model loading.
    """
    img = cv2.imread(image_path)
    if img is None:
        print(f"    ERROR: Cannot read {image_path}")
        return []
    H, W = img.shape[:2]
    all_b, all_s = [], []

    # Full image prediction
    results = yolo_model.predict(image_path, conf=CONFIDENCE_THRES, verbose=False)
    for r in results:
        if r.boxes is not None:
            for box in r.boxes:
                x1, y1, x2, y2 = map(int, box.xyxy[0].tolist())
                all_b.append((x1, y1, x2, y2))
                all_s.append(float(box.conf[0]))

    # Sliced predictions at multiple scales
    for sz in SLICE_SIZES:
        step = int(sz * (1 - OVERLAP_RATIO))
        for y0 in range(0, max(1, H - sz // 4), step):
            for x0 in range(0, max(1, W - sz // 4), step):
                y1c = min(y0 + sz, H)
                x1c = min(x0 + sz, W)
                tile = img[y0:y1c, x0:x1c]
                if tile.shape[0] < 32 or tile.shape[1] < 32:
                    continue
                results = yolo_model.predict(tile, conf=CONFIDENCE_THRES, verbose=False)
                for r in results:
                    if r.boxes is not None:
                        for box in r.boxes:
                            bx1, by1, bx2, by2 = map(int, box.xyxy[0].tolist())
                            # Map back to full image coordinates
                            all_b.append((bx1 + x0, by1 + y0,
                                          bx2 + x0, by2 + y0))
                            all_s.append(float(box.conf[0]))

    # WBF merge across all scales
    mb, ms = wbf([all_b], [all_s])
    recs = []
    for (x1, y1, x2, y2), sc in zip(mb, ms):
        w, h = x2-x1, y2-y1
        if w*h < SYM_MIN_AREA or w < SYM_MIN_DIM or h < SYM_MIN_DIM or w*h > SYM_MAX_AREA:
            continue
        recs.append(dict(x1=x1, y1=y1, x2=x2, y2=y2, width=w, height=h,
                         cx=(x1+x2)//2, cy=(y1+y2)//2,
                         confidence=round(sc, 3), source="yolo_world"))
    print(f"    YOLO-World: {len(recs)} detections")
    return recs


def run_yolo_sahi_fallback(image_path, sahi_model):
    """Fallback: original SAHI-based YOLO detection (for non-World models)."""
    from sahi.predict import get_sliced_prediction
    all_b, all_s = [], []
    for sz in SLICE_SIZES:
        r = get_sliced_prediction(image_path, sahi_model,
            slice_height=sz, slice_width=sz,
            overlap_height_ratio=OVERLAP_RATIO, overlap_width_ratio=OVERLAP_RATIO,
            perform_standard_pred=True, postprocess_type="NMS",
            postprocess_match_metric="IOU", postprocess_match_threshold=NMS_IOU_THRESH)
        b = []; s = []
        for obj in r.object_prediction_list:
            bx = obj.bbox
            b.append((int(bx.minx), int(bx.miny), int(bx.maxx), int(bx.maxy)))
            s.append(obj.score.value)
        all_b.append(b); all_s.append(s)
    mb, ms = wbf(all_b, all_s)
    recs = []
    for (x1, y1, x2, y2), sc in zip(mb, ms):
        w, h = x2-x1, y2-y1
        if w*h < SYM_MIN_AREA or w < SYM_MIN_DIM or h < SYM_MIN_DIM or w*h > SYM_MAX_AREA:
            continue
        recs.append(dict(x1=x1, y1=y1, x2=x2, y2=y2, width=w, height=h,
                         cx=(x1+x2)//2, cy=(y1+y2)//2,
                         confidence=round(sc, 3), source="yolo"))
    print(f"    YOLO/SAHI: {len(recs)} detections")
    return recs


CONTOUR_PASSES = [
    # fine: small symbols (valves, instruments, etc.)
    {"min_area": 80, "max_area": 8000, "solidity": 0.10, "name": "fine"},
    # medium: larger equipment symbols
    {"min_area": 8000, "max_area": 320000, "solidity": 0.08, "name": "medium"},
]

# ── Vector PDF symbol clustering config ────────────────────────────────
VEC_CLUSTER_DIST_PT = 5.0    # pts: max gap between paths belonging to same symbol
VEC_MAX_SYM_W_PT    = 280    # pts: maximum symbol cluster width
VEC_MAX_SYM_H_PT    = 200    # pts: maximum symbol cluster height
VEC_MIN_SYM_AREA_PT = 15     # pt²: minimum cluster bounding-box area
VEC_MIN_ELEMENTS    = 1      # minimum drawing elements per cluster


def extract_vector_symbols_from_pdf(pdf_path, page_num, scale):
    """
    Extract symbol locations from PDF vector graphics (paths/strokes).

    Groups spatially nearby path elements (within VEC_CLUSTER_DIST_PT pts) into
    symbol clusters, filters out page borders and long divider lines, then
    returns detections scaled to the rendered image coordinate space.

    This completely avoids the need for YOLO or image-based contour detection
    when the PDF contains vector symbols (which most CAD exports do).
    """
    doc = fitz.open(pdf_path)
    pg  = doc[page_num - 1]
    drawings = pg.get_drawings()
    doc.close()

    if not drawings:
        print("    No vector drawings in PDF — skipping vector extraction")
        return []

    # ── Filter out page/section borders and ruler-thin dividers ────────
    raw = []
    for d in drawings:
        r = d["rect"]
        w, h = r.width, r.height
        if w <= 0 or h <= 0:
            continue
        # Skip very large rectangles (page frame, section boxes)
        if w > VEC_MAX_SYM_W_PT * 1.5 or h > VEC_MAX_SYM_H_PT * 1.5:
            continue
        # Skip extremely elongated thin lines (column/row dividers)
        short = min(w, h); long = max(w, h)
        if long > 60 and short / long < 0.04:
            continue
        raw.append({"x0": r.x0, "y0": r.y0, "x1": r.x1, "y1": r.y1})

    if not raw:
        print("    All drawings filtered as borders/dividers")
        return []

    # ── Union-Find proximity clustering ────────────────────────────────
    n = len(raw)
    parent = list(range(n))

    def _find(i):
        while parent[i] != i:
            parent[i] = parent[parent[i]]
            i = parent[i]
        return i

    def _union(i, j):
        parent[_find(i)] = _find(j)

    # Sort by y0 for efficient sweep (break early when gap > threshold)
    order = sorted(range(n), key=lambda i: raw[i]["y0"])
    D = VEC_CLUSTER_DIST_PT

    for pi, ii in enumerate(order):
        a = raw[ii]
        for pj in range(pi + 1, n):
            jj = order[pj]
            b = raw[jj]
            # Early exit: b is too far below a even before any horizontal gap
            if b["y0"] - a["y1"] > D + 2:
                break
            # Gap between bounding rectangles (0 if overlapping)
            gap_x = max(0.0, max(a["x0"], b["x0"]) - min(a["x1"], b["x1"]))
            gap_y = max(0.0, max(a["y0"], b["y0"]) - min(a["y1"], b["y1"]))
            dist  = math.hypot(gap_x, gap_y)
            if dist <= D:
                _union(ii, jj)

    # ── Collect clusters and compute bounding boxes ─────────────────────
    from collections import defaultdict
    clusters = defaultdict(list)
    for i in range(n):
        clusters[_find(i)].append(i)

    detections = []
    for members in clusters.values():
        if len(members) < VEC_MIN_ELEMENTS:
            continue
        x0 = min(raw[i]["x0"] for i in members)
        y0 = min(raw[i]["y0"] for i in members)
        x1 = max(raw[i]["x1"] for i in members)
        y1 = max(raw[i]["y1"] for i in members)

        w_pts = x1 - x0
        h_pts = y1 - y0

        if w_pts > VEC_MAX_SYM_W_PT or h_pts > VEC_MAX_SYM_H_PT:
            continue
        if w_pts * h_pts < VEC_MIN_SYM_AREA_PT:
            continue

        # Scale PDF pts → image pixels
        ix1, iy1 = int(x0 * scale), int(y0 * scale)
        ix2, iy2 = int(x1 * scale), int(y1 * scale)
        ws, hs = ix2 - ix1, iy2 - iy1

        if ws < SYM_MIN_DIM or hs < SYM_MIN_DIM:
            continue

        detections.append(dict(
            x1=ix1, y1=iy1, x2=ix2, y2=iy2,
            width=ws, height=hs,
            cx=(ix1 + ix2) // 2, cy=(iy1 + iy2) // 2,
            confidence=0.80,   # vector data = high-confidence detection
            source="vector_pdf",
            n_paths=len(members),
        ))

    print(f"    Vector PDF: {len(detections)} symbol clusters "
          f"(from {len(raw)} filtered paths, {len(drawings)} total)")
    return detections

def run_contours(image_path, existing):
    gray = cv2.imread(image_path, cv2.IMREAD_GRAYSCALE)
    if gray is None:
        print(f"    Contours: cannot read {image_path}")
        return []
    cov = np.zeros(gray.shape, dtype=np.uint8)
    for b in existing:
        cov[b["y1"]:b["y2"], b["x1"]:b["x2"]] = 255
    new = []
    for cfg in CONTOUR_PASSES:
        if cfg["name"] == "fine":
            edges = cv2.Canny(cv2.GaussianBlur(gray, (3, 3), 0), 30, 100)
            k = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (3, 3))
            closed = cv2.morphologyEx(edges, cv2.MORPH_CLOSE, k, iterations=2)
            # Also include adaptive threshold hits for fine symbols
            at = cv2.adaptiveThreshold(cv2.GaussianBlur(gray, (3, 3), 0), 255,
                cv2.ADAPTIVE_THRESH_MEAN_C, cv2.THRESH_BINARY_INV, 11, 3)
            closed = cv2.bitwise_or(closed, at)
        else:
            thresh = cv2.adaptiveThreshold(cv2.GaussianBlur(gray, (5, 5), 0), 255,
                cv2.ADAPTIVE_THRESH_MEAN_C, cv2.THRESH_BINARY_INV, 21, 4)
            k = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (7, 7))
            closed = cv2.morphologyEx(thresh, cv2.MORPH_CLOSE, k, iterations=3)
        # RETR_LIST finds ALL contours (inner + outer), not just outermost.
        # RETR_EXTERNAL only returns the page border when the image has a border,
        # hiding all inner symbol contours.
        cnts, _ = cv2.findContours(closed, cv2.RETR_LIST, cv2.CHAIN_APPROX_SIMPLE)
        for cnt in cnts:
            area = cv2.contourArea(cnt)
            if not(cfg["min_area"] < area < cfg["max_area"]): continue
            ha = cv2.contourArea(cv2.convexHull(cnt))
            if ha > 0 and area / ha < cfg["solidity"]: continue
            x, y, w, h = cv2.boundingRect(cnt)
            if w < SYM_MIN_DIM or h < SYM_MIN_DIM: continue
            cx, cy = x + w//2, y + h//2
            if cov[cy, cx] > 0: continue
            new.append(dict(x1=x, y1=y, x2=x+w, y2=y+h, width=w, height=h,
                            cx=cx, cy=cy, confidence=0.22,
                            source=f"contour_{cfg['name']}"))
            cov[y:y+h, x:x+w] = 255
    print(f"    Contours: {len(new)} additional")
    return new

def nms_all(dets, iou_thr=0.45):
    if not dets: return []
    boxes = [(d["x1"], d["y1"], d["x2"], d["y2"]) for d in dets]
    scores = [d["confidence"] for d in dets]
    order = sorted(range(len(scores)), key=lambda i: -scores[i])
    keep = []; used = [False]*len(dets)
    for i in order:
        if used[i]: continue
        keep.append(i); used[i] = True
        for j in order:
            if used[j]: continue
            if iou(boxes[i], boxes[j]) > iou_thr:
                used[j] = True
    print(f"    After NMS: {len(keep)} symbols")
    return [dets[i] for i in keep]


def get_nearby_ocr(sym, classified_groups):
    """Collect ALL text within a generous radius of the symbol."""
    sx1, sy1, sx2, sy2 = sym["x1"], sym["y1"], sym["x2"], sym["y2"]
    sw, sh = sym["width"], sym["height"]
    pad = max(80, int(max(sw, sh) * 1.2))
    nearby = []
    for grp in classified_groups:
        if grp.get("kind") == "header": continue
        if (grp["x1"] <= sx2 + pad and grp["x2"] >= sx1 - pad and
                grp["y1"] <= sy2 + pad and grp["y2"] >= sy1 - pad):
            nearby.append(grp["text"])
    return " ".join(nearby)


# ═══════════════════════════════════════════════════════════════════════
# FIX #5: Omnidirectional section assignment
# Handles headers to the right, left, above, and below the symbol.
# Uses containment and nearest-distance across all directions.
# ═══════════════════════════════════════════════════════════════════════
def assign_sections(symbols, classified_groups):
    """
    Column-aware section assignment for grid-layout legend sheets.

    Priority 1: headers that are ABOVE the symbol (header_cy < symbol_cy)
                AND within COLUMN_WIDTH_PX horizontally (same column).
                Among those, pick the nearest.
    Priority 2: fallback to nearest header overall (original behaviour)
                when no column-above header exists.

    This prevents the page-bottom SWITCHING CONTROLS header from stealing
    symbols that belong to the LIGHTING section above it in the same column.
    """
    headers = [g for g in classified_groups if g.get("kind") == "header"]
    if not headers:
        for sym in symbols:
            sym["section"] = "General"
        return symbols

    # Roughly one column width in pixels at the current DPI.
    # Columns are ~396 pts apart; at 150 DPI that is ~825 px.
    # Use 800 px so cross-column headers are excluded but same-column ones
    # (even slightly off-center) are included.
    COLUMN_WIDTH_PX = 800

    for sym in symbols:
        scx, scy = sym["cx"], sym["cy"]
        best_hdr = "General"
        best_dist = float("inf")

        # Pass 1: same column, above symbol
        for hdr in headers:
            if hdr["cy"] >= scy:
                continue
            if abs(hdr["cx"] - scx) > COLUMN_WIDTH_PX:
                continue
            dist = math.hypot(hdr["cx"] - scx, hdr["cy"] - scy)
            if dist < best_dist:
                best_dist = dist
                best_hdr = hdr["text"]

        # Pass 2: any direction, any column (fallback for symbols at very top)
        if best_hdr == "General":
            for hdr in headers:
                dist = math.hypot(hdr["cx"] - scx, hdr["cy"] - scy)
                if dist < best_dist:
                    best_dist = dist
                    best_hdr = hdr["text"]

        sym["section"] = best_hdr
    return symbols


# ═══════════════════════════════════════════════════════════════════════
# FIX #12: Grid-aware symbol↔label pairing
# ═══════════════════════════════════════════════════════════════════════
def pair_symbols_with_labels(symbols, classified_groups):
    """
    FIX #12: For legend sheets, symbols and their labels form a grid:
    - Symbol on the LEFT, label text on the RIGHT (same row)
    - Within each section, pair symbols 1:1 with nearest right-side label

    This replaces the pure-distance approach with a structured grid match.
    """
    # Exclude sub-column header words that appear just below section headers
    # (e.g. "SYMBOL" / "DESCRIPTION" at Y≈241 in each column of the legend).
    _SUBLABEL_SKIP = {"SYMBOL", "DESCRIPTION"}
    labels = [g for g in classified_groups
              if g.get("kind") == "label"
              and g["text"].strip().upper() not in _SUBLABEL_SKIP]
    if not labels or not symbols:
        return symbols

    used_labels = set()

    for sym in symbols:
        sx1, sy1, sx2, sy2 = sym["x1"], sym["y1"], sym["x2"], sym["y2"]
        scy = sym["cy"]

        best_label = None
        best_score = float("inf")

        for li, lbl in enumerate(labels):
            if li in used_labels:
                continue

            lx1, ly1, lx2, ly2 = lbl["x1"], lbl["y1"], lbl["x2"], lbl["y2"]
            lcy = lbl["cy"]

            # Label must be on the same horizontal band.
            # Use ±60px to accommodate multi-line labels and font-height offsets.
            if abs(scy - lcy) > 60:
                continue

            # Prefer labels to the RIGHT of the symbol.
            # Use symbol cx (not x2) so over-wide clusters still pair correctly.
            dx = lx1 - sym["cx"]
            if dx < -10:
                continue  # label clearly left of symbol centre
            if dx > ASSOC_RIGHT_MAX:
                continue  # too far right

            # Score: prefer same-row and close horizontal
            score = abs(scy - lcy) * 2 + abs(dx)
            if score < best_score:
                best_score = score
                best_label = (li, lbl)

        if best_label is not None:
            li, lbl = best_label
            used_labels.add(li)
            # Collect continuation lines: labels in the same x-column that
            # sit just below the first-matched label (each row ≈ 20-35 px).
            # This assembles multi-line descriptions like:
            #   "FIRE ALARM CONTROL PANEL AND ASSOCIATED"  ← first line
            #   "COMPONENTS."                              ← continuation
            first_lcy = lbl["cy"]
            first_lx1 = lbl["x1"]
            continuation_lines = [lbl["text"]]
            for li2, lbl2 in enumerate(labels):
                if li2 in used_labels or li2 == li:
                    continue
                dy = lbl2["cy"] - first_lcy
                # Must be below (dy > 0) and within 1 line-height (~28 px).
                # Genuine wrapping continuation lines are 12-20 px below;
                # next-symbol labels are 30-80 px below.
                if not (0 < dy <= 28):
                    continue
                # Must overlap the same x-band as the first label
                if lbl2["x2"] < first_lx1 - 20 or lbl2["x1"] > lbl["x2"] + 20:
                    continue
                continuation_lines.append(lbl2["text"])
                used_labels.add(li2)
            sym["paired_label"] = " ".join(continuation_lines)
        else:
            # Fallback: nearest label in any direction (within range)
            for li, lbl in enumerate(labels):
                if li in used_labels:
                    continue
                dist = math.hypot(lbl["cx"] - sym["cx"], lbl["cy"] - scy)
                if dist < ASSOC_ANY_MAX and dist < best_score:
                    best_score = dist
                    best_label = (li, lbl)
            if best_label is not None:
                li, lbl = best_label
                used_labels.add(li)
                sym["paired_label"] = lbl["text"]
            else:
                sym["paired_label"] = ""

    return symbols


# ═══════════════════════════════════════════════════════════════════════
# ROW-ANCHORED LEGEND EXTRACTION
# ═══════════════════════════════════════════════════════════════════════
def extract_row_crops(classified_groups, color_img):
    """
    For each description label in the legend, crop the symbol column at that
    row's Y extent.  Section assignment runs before continuation assembly so
    that labels from adjacent columns never contaminate each other's grouping.
    """
    H, W = color_img.shape[:2]
    _SKIP = {"SYMBOL", "DESCRIPTION"}

    # "DESCRIPTION" sub-header texts reliably mark each column's description-
    # start X.  Collect ALL occurrences regardless of classification kind
    # (they may be labelled as header, label, or noise depending on font size).
    desc_col_anchors = sorted(
        [g for g in classified_groups
         if g["text"].strip().upper() == "DESCRIPTION"],
        key=lambda g: g["cx"],
    )

    raw_labels = [g for g in classified_groups
                  if g.get("kind") == "label"
                  and g["text"].strip().upper() not in _SKIP]

    # ── Step 1: Assign sections to raw labels FIRST ───────────────────────
    raw_labels = assign_sections(raw_labels, classified_groups)

    # ── Step 2: Assemble multi-line descriptions per section ─────────────
    by_section_raw = defaultdict(list)
    for lbl in raw_labels:
        by_section_raw[lbl.get("section", "General")].append(lbl)

    assembled_all = []
    for sec_name, sec_raw in by_section_raw.items():
        sec_sorted_raw = sorted(sec_raw, key=lambda l: l["cy"])
        used = set()
        for i, lbl in enumerate(sec_sorted_raw):
            if id(lbl) in used:
                continue
            used.add(id(lbl))
            lines    = [lbl["text"]]
            acc_y2   = lbl["y2"]
            acc_x1   = lbl["x1"]
            acc_x2   = lbl["x2"]
            first_cy = lbl["cy"]
            for lbl2 in sec_sorted_raw[i + 1:]:
                if id(lbl2) in used:
                    continue
                dy = lbl2["cy"] - first_cy
                if dy > 40:
                    break
                if dy < 0 or dy > 28:
                    continue
                # (a) Horizontal fragment — lbl2 starts to the right of acc_x2
                # (b) Vertical continuation — lbl2 X-range overlaps the group
                if lbl2["x1"] < acc_x2 - 20:
                    x_ov = (min(acc_x2, lbl2["x2"])
                            - max(acc_x1, lbl2["x1"]))
                    if x_ov <= 0:
                        continue
                lines.append(lbl2["text"])
                acc_y2 = max(acc_y2, lbl2["y2"])
                acc_x1 = min(acc_x1, lbl2["x1"])
                acc_x2 = max(acc_x2, lbl2["x2"])
                used.add(id(lbl2))
            assembled_all.append({
                **lbl,
                "text":    " ".join(lines),
                "y2":      acc_y2,
                "x1":      acc_x1,
                "x2":      acc_x2,
                "cy":      (lbl["y1"] + acc_y2) // 2,
                "section": sec_name,
            })

    # ── Step 3: Build one crop per assembled row ──────────────────────────
    by_section = defaultdict(list)
    for lbl in assembled_all:
        by_section[lbl.get("section", "General")].append(lbl)

    headers = {g["text"]: g for g in classified_groups if g.get("kind") == "header"}
    symbols = []

    # Maximum width of a symbol cell at 150 DPI (~400px = ~2.7 inches)
    SYM_COL_WIDTH = 400

    for sec_name, sec_labels in by_section.items():
        sec_sorted = sorted(sec_labels, key=lambda l: l["cy"])
        sec_hdr    = headers.get(sec_name)

        for i, lbl in enumerate(sec_sorted):
            # Crop right edge = just before this label's description text.
            # lbl["x1"] is exactly where the description text starts for this row,
            # so this is always correct regardless of which column the label is in.
            sym_x2 = max(0, lbl["x1"] - 10)
            sym_x1 = max(0, lbl["x1"] - SYM_COL_WIDTH)
            if sym_x2 <= sym_x1:
                continue

            # Row Y: midpoints between adjacent assembled labels
            if i == 0:
                # Guard against sec_hdr pointing to a header below the first label
                top_ref = (int(sec_hdr["y2"])
                           if sec_hdr and sec_hdr["y2"] < lbl["y1"]
                           else lbl["y1"])
                row_y1  = max(0, (top_ref + lbl["y1"]) // 2)
            else:
                row_y1 = max(0, (sec_sorted[i - 1]["y2"] + lbl["y1"]) // 2)

            if i == len(sec_sorted) - 1:
                row_y2 = min(H, lbl["y2"] + 20)
            else:
                row_y2 = min(H, (lbl["y2"] + sec_sorted[i + 1]["y1"]) // 2)

            if row_y2 <= row_y1:
                continue

            crop = color_img[row_y1:row_y2, sym_x1:sym_x2].copy()
            symbols.append({
                "section":      sec_name,
                "paired_label": lbl["text"],
                "image_color":  crop if crop.size > 0 else None,
                "x1": sym_x1, "y1": row_y1, "x2": sym_x2, "y2": row_y2,
                "cx": (sym_x1 + sym_x2) // 2,
                "cy": (row_y1 + row_y2) // 2,
                "width":  sym_x2 - sym_x1,
                "height": row_y2 - row_y1,
                "confidence": 1.0,
                "source":     "legend_row",
            })

    print(f"    Row crops: {len(symbols)} rows from {len(assembled_all)} labels")
    return symbols


# ═══════════════════════════════════════════════════════════════════════
# EXCEL OUTPUT  (professional 4-sheet workbook)
# ═══════════════════════════════════════════════════════════════════════
_HDR_FILL  = PatternFill("solid", fgColor="0D2137")
_HDR_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
_HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_THIN      = Side(style="thin", color="BBBBBB")
_BDR       = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_ALT       = PatternFill("solid", fgColor="EDF2FA")
_WHITE     = PatternFill("solid", fgColor="FFFFFF")

_SEC_COLORS = {
    "LINE SYMBOLS":           "D6E4F7",
    "PIPING SYMBOLS":         "D6F0D6",
    "CONTROL VALVE SYMBOLS":  "FAF0D6",
    "VALVE SYMBOLS":          "FAF0D6",
    "INSTRUMENT SYMBOLS":     "F7D6F0",
    "INSULATION SYMBOLS":     "F0F7D6",
    "VESSELS":                "D6F7F0",
    "PUMPS AND COMPRESSOR":   "F7E6D6",
    "COLUMN AND HEAT EXCHANGERS": "E6D6F7",
    "MISCELLANEOUS ITEMS":    "F7F0D6",
}

def _hrow(ws, row, vals, widths):
    ws.row_dimensions[row].height = 30
    for ci, (v, w) in enumerate(zip(vals, widths), 1):
        c = ws.cell(row=row, column=ci, value=v)
        c.fill = _HDR_FILL; c.font = _HDR_FONT
        c.alignment = _HDR_ALIGN; c.border = _BDR
        ws.column_dimensions[get_column_letter(ci)].width = w

def _dcell(ws, row, col, val, fill, bold=False):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = fill; c.border = _BDR
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if bold: c.font = Font(name="Arial", bold=True, size=10)
    return c

def _thumb(ws, addr, crop):
    if crop is None or crop.size == 0: return
    try:
        t = cv2.resize(crop, (THUMB_W, THUMB_H))
        rgb = cv2.cvtColor(t, cv2.COLOR_BGR2RGB) if t.ndim == 3 else t
        buf = io.BytesIO(); PILImage.fromarray(rgb).save(buf, format="PNG"); buf.seek(0)
        xl = XLImage(buf); xl.width = THUMB_W; xl.height = THUMB_H
        ws.add_image(xl, addr)
    except Exception: pass

def _sec_fill(section):
    sec_up = section.upper()
    for k, v in _SEC_COLORS.items():
        if k in sec_up: return PatternFill("solid", fgColor=v)
    return PatternFill("solid", fgColor="F5F5F5")


def save_excel(all_symbols, all_unmatched, all_classified_flat, ocr_tokens_by_page, out_dir):
    """
    FIX #11: Receives separate classified_groups (flat list) AND
    ocr_tokens_by_page (dict). Previously both were the same dict.
    """
    print("\n[EXCEL] Writing legend_symbols.xlsx …")
    wb = Workbook()

    # ── Sheet 1: All Symbols ────────────────────────────────────
    ws = wb.active; ws.title = "All Symbols"
    COLS = ["#", "Page", "Section", "Symbol Name (Confirmed)", "Standard Ref",
          "VLM Confidence", "VLM Reasoning", "Nearby OCR Text", "Paired Label",
          "YOLO Conf", "Source", "X1", "Y1", "X2", "Y2", "W px", "H px",
          "Symbol Image"]
    WDS = [5, 5, 22, 30, 14, 11, 35, 30, 25, 9, 12, 7, 7, 7, 7, 7, 7, 13]
    _hrow(ws, 1, COLS, WDS)

    all_symbols.sort(key=lambda s: (s.get("page", 1), s.get("section", "ZZZ"),
                                    s.get("cy", 0), s.get("cx", 0)))
    cur_sec = None; row_n = 2
    for sym in all_symbols:
        sec = sym.get("section", "General")
        if sec != cur_sec:
            cur_sec = sec
            sf = _sec_fill(sec)
            for ci in range(1, len(COLS)+1):
                c = ws.cell(row=row_n, column=ci); c.fill = sf; c.border = _BDR
            ws.cell(row=row_n, column=1, value="▶").alignment = _HDR_ALIGN
            c2 = ws.cell(row=row_n, column=3, value=sec)
            c2.font = Font(name="Arial", bold=True, size=10)
            c2.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[row_n].height = 16; row_n += 1

        fill = _ALT if row_n % 2 == 0 else _WHITE
        ws.row_dimensions[row_n].height = THUMB_H * 0.72
        vlm = sym.get("vlm", {})
        vals = [sym.get("item_id", ""), sym.get("page", ""), sec,
              vlm.get("name", "Unknown"), vlm.get("standard_ref", ""),
              vlm.get("confidence", 0), vlm.get("reasoning", ""),
              sym.get("ocr_nearby", ""), sym.get("paired_label", ""),
              sym.get("confidence", 0), sym.get("source", ""),
              sym["x1"], sym["y1"], sym["x2"], sym["y2"],
              sym["width"], sym["height"]]
        for ci, v in enumerate(vals, 1): _dcell(ws, row_n, ci, v, fill)
        # Confidence colour
        conf = vlm.get("confidence", 0)
        conf_fill = (PatternFill("solid", fgColor="C6EFCE") if conf >= 0.7 else
                   PatternFill("solid", fgColor="FFEB9C") if conf >= 0.4 else
                   PatternFill("solid", fgColor="FFC7CE"))
        ws.cell(row=row_n, column=6).fill = conf_fill
        _thumb(ws, f"R{row_n}", sym.get("image_color"))
        row_n += 1

    # ── Sheet 2: Section Summary ─────────────────────────────────
    ws2 = wb.create_sheet("Section Summary")
    _hrow(ws2, 1, ["Section", "Total", "High Conf (≥0.7)", "Med Conf", "Low Conf",
                    "YOLO-World", "Contour Fine", "Contour Med"],
          [28, 8, 13, 10, 10, 12, 12, 12])
    sec_data = defaultdict(list)
    for s in all_symbols: sec_data[s.get("section", "General")].append(s)
    for ri, (sec, syms) in enumerate(sorted(sec_data.items()), 2):
        fill = _ALT if ri % 2 == 0 else _WHITE
        confs = [s.get("vlm", {}).get("confidence", 0) for s in syms]
        hi = sum(1 for c in confs if c >= 0.7)
        med = sum(1 for c in confs if 0.4 <= c < 0.7)
        lo = sum(1 for c in confs if c < 0.4)
        vals = [sec, len(syms), hi, med, lo,
              sum(1 for s in syms if "yolo" in s.get("source", "")),
              sum(1 for s in syms if s.get("source") == "contour_fine"),
              sum(1 for s in syms if s.get("source") == "contour_medium")]
        for ci, v in enumerate(vals, 1): _dcell(ws2, ri, ci, v, fill)
    if sec_data:
        chart = BarChart(); chart.type = "col"; chart.title = "Symbols per Section"
        chart.y_axis.title = "Count"
        data = Reference(ws2, min_col=2, min_row=1, max_row=len(sec_data)+1)
        cats = Reference(ws2, min_col=1, min_row=2, max_row=len(sec_data)+1)
        chart.add_data(data, titles_from_data=True); chart.set_categories(cats)
        chart.shape = 4; chart.width = 26; chart.height = 14
        ws2.add_chart(chart, "J2")

    # ── Sheet 3: Raw OCR Text (per page) ─────────────────────────
    ws3 = wb.create_sheet("OCR Raw Text")
    _hrow(ws3, 1, ["#", "Page", "Text", "Kind", "Confidence", "Tokens Merged",
                  "X1", "Y1", "X2", "Y2"],
          [5, 5, 45, 8, 10, 10, 7, 7, 7, 7])
    ri = 2
    for page, classified in sorted(ocr_tokens_by_page.items()):
        for grp in sorted(classified, key=lambda g: (g.get("y1", 0))):
            fill = _ALT if ri % 2 == 0 else _WHITE
            kind = grp.get("kind", "?")
            kind_fill = (PatternFill("solid", fgColor="D6E4F7") if kind == "header" else
                       PatternFill("solid", fgColor="D6F0D6") if kind == "label" else
                       PatternFill("solid", fgColor="F5F5F5"))
            vals = [ri-1, page, grp["text"], kind, grp["confidence"],
                  grp.get("token_count", 1),
                  grp["x1"], grp["y1"], grp["x2"], grp["y2"]]
            for ci, v in enumerate(vals, 1):
                c = _dcell(ws3, ri, ci, v, fill)
                if ci == 4: c.fill = kind_fill
            ri += 1

    # ── Sheet 4: Unmatched Labels ─────────────────────────────────
    ws4 = wb.create_sheet("Unmatched Labels")
    _hrow(ws4, 1, ["#", "Page", "Text", "Confidence", "X1", "Y1", "X2", "Y2"],
          [5, 5, 45, 10, 7, 7, 7, 7])
    for ri, grp in enumerate(all_unmatched, 2):
        fill = _ALT if ri % 2 == 0 else _WHITE
        vals = [ri-1, grp.get("page", "?"), grp["text"], grp["confidence"],
              grp["x1"], grp["y1"], grp["x2"], grp["y2"]]
        for ci, v in enumerate(vals, 1): _dcell(ws4, ri, ci, v, fill)

    out = out_dir / "legend_symbols.xlsx"
    wb.save(str(out)); print(f"    → Saved: {out}")
    return out


# ═══════════════════════════════════════════════════════════════════════
# DEBUG ANNOTATED IMAGE
# ═══════════════════════════════════════════════════════════════════════
def draw_debug(symbols, classified, orig_path, out_dir, page):
    img = cv2.imread(orig_path)
    if img is None: return
    H, W = img.shape[:2]
    scale = max(0.22, min(0.60, W/5500)); thick = max(1, int(scale*2))
    for grp in classified:
        col = (255, 100, 0) if grp.get("kind") == "header" else (0, 180, 0)
        cv2.rectangle(img, (grp["x1"], grp["y1"]), (grp["x2"], grp["y2"]), col, 1)
    for sym in symbols:
        vlm = sym.get("vlm", {}); conf = vlm.get("confidence", 0)
        col = ((0, 200, 60) if conf >= 0.7 else
               (0, 180, 220) if conf >= 0.4 else (0, 80, 220))
        x1, y1, x2, y2 = sym["x1"], sym["y1"], sym["x2"], sym["y2"]
        cv2.rectangle(img, (x1, y1), (x2, y2), col, thick+1)
        label = (vlm.get("name", "?")[:22] + f" {conf:.2f}")
        (tw, th), _ = cv2.getTextSize(label, cv2.FONT_HERSHEY_SIMPLEX, scale, thick)
        cv2.rectangle(img, (x1, y1-th-5), (x1+tw+4, y1), col, cv2.FILLED)
        cv2.putText(img, label, (x1+2, y1-4), cv2.FONT_HERSHEY_SIMPLEX,
                    scale, (255, 255, 255), thick, cv2.LINE_AA)
    out = out_dir / f"debug_v7_p{page}.png"
    cv2.imwrite(str(out), img)
    print(f"    Debug → {out.name}")


# ═══════════════════════════════════════════════════════════════════════
# FIX #7: Improved unmatched label detection
# ═══════════════════════════════════════════════════════════════════════
def find_unmatched_labels(symbols, usable_labels):
    """
    FIX #7: Compare label text against:
    1. VLM-confirmed names (canonical)
    2. Alias lookup (maps label text → canonical name)
    3. Paired label text from grid matching
    4. Nearby OCR text collected for each symbol

    Previous version compared raw OCR text against VLM canonical names
    which never matched (e.g. "gate" ≠ "Gate Valve").
    """
    # Build set of all matched identifiers
    matched = set()
    for s in symbols:
        vlm_name = s.get("vlm", {}).get("name", "").lower()
        if vlm_name:
            matched.add(vlm_name)
        paired = s.get("paired_label", "").lower()
        if paired:
            matched.add(paired)
        # Also add nearby OCR text fragments
        for fragment in s.get("ocr_nearby", "").lower().split():
            if len(fragment) >= 3:
                matched.add(fragment)

    unmatched = []
    for g in usable_labels:
        label_lower = g["text"].lower().strip()
        # Check 1: exact match in matched set
        if label_lower in matched:
            continue
        # Check 2: any word from the label appears in matched
        label_words = set(w for w in label_lower.split() if len(w) >= 3)
        if label_words and label_words & matched:
            continue
        # Check 3: alias lookup → canonical name found in matched
        canonical = _ALIAS_LOOKUP.get(label_lower, "").lower()
        if canonical and canonical in matched:
            continue
        # Check 4: label is a substring of any matched text
        if any(label_lower in m for m in matched if len(m) >= 3):
            continue
        # None of the checks matched → truly unmatched
        unmatched.append(g)

    return unmatched


# ═══════════════════════════════════════════════════════════════════════
# CLI & RUNTIME CONFIG
# ═══════════════════════════════════════════════════════════════════════
def _safe_filename(text, max_len=120):
    """Sanitize a string so it can be used as a file/directory name."""
    s = re.sub(r'[\\/:*?"<>|]', '_', (text or "").strip())
    s = re.sub(r'\s+', ' ', s).strip()
    return s[:max_len] if s else "UNLABELED"


def _parse_args():
    p = argparse.ArgumentParser(
        description="P&ID Legend Analyzer v7 — extract symbols from legend PDFs"
    )
    p.add_argument("-i", "--input",    required=True, type=Path,
                   help="Path to input Legend PDF file")
    p.add_argument("-o", "--output",   required=True, type=Path,
                   help="Directory for extracted assets")
    p.add_argument("-db", "--database", required=True, type=Path,
                   help="Path to symbols JSON knowledge-base file")
    p.add_argument("-p", "--pages",    default="all",
                   help="Pages to process: 'all' or comma-separated (e.g. '1,2,5')")
    p.add_argument("--vlm",  action="store_true",
                   help="Enable VLM confirmation (requires OPENAI_API_KEY env var)")
    p.add_argument("--clip", action="store_true",
                   help="Enable CLIP embedding matching (requires transformers)")
    return p.parse_args()


def _load_config(args):
    """Load knowledge-base and hyperparameters from JSON; apply CLI flag overrides."""
    global PID_KNOWLEDGE_BASE, PID_YOLO_CLASSES
    global PDF_DPI, YOLO_MODEL_PATH
    global H_GAP_PX, MAX_GROUP_WIDTH_PX, CONFIDENCE_THRES
    global CLIP_MODEL_NAME, VLM_MODEL, USE_VLM, USE_CLIP
    global _ALIAS_ENTRIES, _ALIAS_LOOKUP

    db_path = args.database.resolve()
    if not db_path.exists():
        print(f"ERROR: database file not found: {db_path}")
        sys.exit(1)

    try:
        with open(db_path, "r", encoding="utf-8") as f:
            db = json.load(f)
    except json.JSONDecodeError as e:
        print(f"ERROR: invalid JSON in database '{db_path}': {e}")
        sys.exit(1)

    if "knowledge_base" not in db:
        print("ERROR: database JSON must contain a 'knowledge_base' array")
        sys.exit(1)

    PID_KNOWLEDGE_BASE[:] = db["knowledge_base"]
    PID_YOLO_CLASSES[:] = db.get("yolo_classes", [])

    hp = db.get("hyperparameters", {})
    PDF_DPI             = hp.get("pdf_dpi",             PDF_DPI)
    YOLO_MODEL_PATH     = hp.get("yolo_model_path",     YOLO_MODEL_PATH)
    H_GAP_PX            = hp.get("h_gap_px",            H_GAP_PX)
    MAX_GROUP_WIDTH_PX  = hp.get("max_group_width_px",  MAX_GROUP_WIDTH_PX)
    CONFIDENCE_THRES    = hp.get("confidence_thres",    CONFIDENCE_THRES)
    CLIP_MODEL_NAME     = hp.get("clip_model_name",     CLIP_MODEL_NAME)
    VLM_MODEL           = hp.get("vlm_model",           VLM_MODEL)

    USE_VLM  = args.vlm
    USE_CLIP = args.clip

    _ALIAS_ENTRIES.clear()
    _ALIAS_LOOKUP.clear()
    for entry in PID_KNOWLEDGE_BASE:
        for alias in ([entry["name"].lower()] +
                      [a.lower() for a in entry.get("aliases", [])]):
            _ALIAS_ENTRIES.append((alias, entry["name"], len(alias)))
            _ALIAS_LOOKUP[alias] = entry["name"]
    _ALIAS_ENTRIES.sort(key=lambda x: -x[2])

    print(f"  [DB] Loaded {len(PID_KNOWLEDGE_BASE)} KB entries, "
          f"{len(PID_YOLO_CLASSES)} YOLO classes from {db_path.name}")
    if USE_VLM:
        print(f"  [VLM] Enabled — model: {VLM_MODEL}")
    if USE_CLIP:
        print(f"  [CLIP] Enabled — model: {CLIP_MODEL_NAME}")


# ═══════════════════════════════════════════════════════════════════════
# MAIN PIPELINE
# ═══════════════════════════════════════════════════════════════════════
def process_page(pdf_path, page, dpi, out_dir, yolo_model, yolo_type, ocr_reader):
    print(f"\n{'─'*58}")
    print(f"  PAGE {page}")
    print(f"{'─'*58}")

    img_path  = pdf_to_image(pdf_path, page, dpi, out_dir, prefix="legend")
    color_img = cv2.imread(img_path)
    H, W      = color_img.shape[:2]
    scale     = dpi / 72.0

    print(f"\n  [1] Text extraction …")
    tokens = run_ocr_raw(img_path, ocr_reader,
                         pdf_path=pdf_path, page_num=page, scale=scale)

    print(f"\n  [2] Token grouping …")
    groups = group_tokens(tokens)
    for g in groups:
        g["page"] = page

    print(f"\n  [3] Classify …")
    classified = classify_text_groups(groups)

    print(f"\n  [4] Row-anchored extraction …")
    all_syms = extract_row_crops(classified, color_img)

    print(f"\n  [5] Save crops + KB ({len(all_syms)} symbols) …")
    legends_dir = out_dir / "legends"
    _used_names: dict = {}

    for idx, sym in enumerate(all_syms):
        sym["page"]       = page
        sym["item_id"]    = f"P{page}_{idx + 1:04d}"
        sym["ocr_nearby"] = sym["paired_label"]

        crop = sym.get("image_color")
        if crop is not None and crop.size > 0:
            sec_dir = legends_dir / _safe_filename(sym.get("section", "General"))
            sec_dir.mkdir(parents=True, exist_ok=True)
            base  = _safe_filename(sym.get("paired_label") or sym["item_id"])
            key   = (str(sec_dir), base)
            count = _used_names.get(key, 0)
            _used_names[key] = count + 1
            fname = f"{base}.png" if count == 0 else f"{base}_{count + 1}.png"
            cv2.imwrite(str(sec_dir / fname), crop)
            sym["legend_path"] = str(sec_dir / fname)

        candidates = kb_retrieve(crop, sym["paired_label"], top_k=VLM_TOP_K)

        vlm_result = vlm_confirm(crop, sym["paired_label"], candidates, sym["section"])
        # When neither CLIP nor VLM is active, use the actual label text as the name
        if not USE_VLM and not USE_CLIP and sym.get("paired_label"):
            vlm_result["name"] = sym["paired_label"]
        sym["vlm"] = vlm_result

        if (idx + 1) % 20 == 0 or idx == len(all_syms) - 1:
            pct = int((idx + 1) / len(all_syms) * 100)
            print(f"    {idx + 1}/{len(all_syms)} ({pct}%)")

    print(f"\n  [6] Debug image …")
    draw_debug(all_syms, classified, img_path, out_dir, page)

    return all_syms, [], classified


def main():
    args = _parse_args()
    _load_config(args)

    print("=" * 60)
    print("  P&ID Legend Analyzer  v7  — All Fixes Applied")
    print("=" * 60)

    out_dir  = setup_dirs(args.output)
    pdf_path = str(args.input.resolve())

    if not args.input.exists():
        print(f"ERROR: input PDF not found: {args.input}")
        sys.exit(1)

    if args.pages.strip().lower() == "all":
        doc   = fitz.open(pdf_path)
        pages = list(range(1, doc.page_count + 1))
        doc.close()
        print(f"  Pages: all ({len(pages)} total)")
    else:
        try:
            pages = [int(p.strip()) for p in args.pages.split(",") if p.strip()]
        except ValueError:
            print("ERROR: --pages must be 'all' or comma-separated integers (e.g. '1,2,5')")
            sys.exit(1)
        if not pages:
            print("ERROR: --pages produced an empty page list")
            sys.exit(1)
        print(f"  Pages: {pages}")

    ocr_reader = None
    if EASYOCR_AVAILABLE:
        try:
            ocr_reader = easyocr.Reader(["en"], gpu=False, verbose=False)
            print("  [OCR] EasyOCR loaded (fallback)")
        except Exception as e:
            print(f"  [OCR] EasyOCR init failed: {e}")
    else:
        print("  [OCR] EasyOCR not installed — will use PDF text layer")

    yolo_model, yolo_type = _load_yolo_world(YOLO_MODEL_PATH)
    build_kb_vectors()

    all_symbols = []; all_unmatched = []; ocr_by_page = {}
    all_classified_flat = []

    for page in pages:
        syms, unmatched, classified = process_page(
            pdf_path, page, PDF_DPI, out_dir,
            yolo_model, yolo_type, ocr_reader)
        all_symbols += syms
        all_unmatched += unmatched
        ocr_by_page[page] = classified
        all_classified_flat.extend(classified)

    print(f"\n{'='*60}")
    print(f"  Total symbols : {len(all_symbols)}")
    print(f"  Unmatched labels: {len(all_unmatched)}")

    save_excel(all_symbols, all_unmatched, all_classified_flat, ocr_by_page, out_dir)

    print(f"\n  Outputs → {args.output}")
    print(f"    legend_symbols.xlsx    4 sheets")
    print(f"    debug_v7_p*.png        colour-coded by confidence")
    print(f"    *_clean.png            text-erased images")
    print(f"    crops/                 per-symbol crops")
    print("=" * 60)

if __name__ == "__main__":
    main()
